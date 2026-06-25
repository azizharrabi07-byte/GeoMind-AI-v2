"""
File Parser Service — Extracts text and metadata from survey file formats.
Supports: PDF, CSV, XLSX, DXF, GeoJSON, Images (OCR), TXT
"""
import io
import logging
import re
from typing import Optional

logger = logging.getLogger("geomind.files")


def parse_file(filename: str, content: bytes) -> dict:
    """Parse an uploaded file and extract text + metadata."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    result = {"ext": ext, "text": "", "metadata": {}, "error": None}

    try:
        if ext == "pdf":
            return _parse_pdf(content)
        elif ext in ("xlsx", "xls"):
            return _parse_xlsx(content)
        elif ext == "csv":
            return _parse_csv(content)
        elif ext == "dxf":
            return _parse_dxf(content)
        elif ext in ("geojson", "json"):
            return _parse_geojson(content)
        elif ext in ("png", "jpg", "jpeg", "tiff", "tif", "bmp", "webp"):
            return _parse_image(content)
        elif ext in ("txt", "raw", "gsi", "sdr", "rw5"):
            return _parse_text(content)
        else:
            result["text"] = content.decode("utf-8", errors="replace")
            result["metadata"]["format"] = ext
            return result
    except Exception as e:
        logger.error(f"Parse error for {filename}: {e}")
        result["text"] = f"[Parse error: {e}]"
        result["error"] = str(e)
        return result


def _parse_pdf(content: bytes) -> dict:
    """Extract text from PDF files."""
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    text = "\n".join([page.extract_text() or "" for page in reader.pages])
    return {
        "ext": "pdf",
        "text": text,
        "metadata": {
            "pages": len(reader.pages),
            "title": reader.metadata.title if reader.metadata else None,
        },
    }


def _parse_xlsx(content: bytes) -> dict:
    """Extract data from Excel files."""
    import pandas as pd
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    texts = []
    total_rows = 0

    for name in sheet_names[:5]:  # Limit to first 5 sheets
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c or "") for c in rows[0]]
            data_rows = rows[1:51]  # First 50 data rows
            texts.append(f"Sheet: {name}")
            texts.append(f"Columns: {', '.join(header)}")
            texts.append(f"Rows: {len(rows) - 1}")
            for row in data_rows:
                texts.append(", ".join([str(c or "") for c in row]))
            total_rows += len(rows)

    return {
        "ext": "xlsx",
        "text": "\n\n".join(texts),
        "metadata": {"sheets": sheet_names, "total_rows": total_rows},
    }


def _parse_csv(content: bytes) -> dict:
    """Extract data from CSV files."""
    text = content.decode("utf-8", errors="replace")
    lines = text.strip().split("\n")
    headers = lines[0].split(",") if lines else []
    return {
        "ext": "csv",
        "text": text,
        "metadata": {
            "headers": [h.strip() for h in headers],
            "row_count": max(0, len(lines) - 1),
        },
    }


def _parse_dxf(content: bytes) -> dict:
    """Extract entities, layers, and coordinates from ASCII DXF."""
    text = content.decode("utf-8", errors="replace")
    lines = text.split("\n")
    layers = set()
    entities = []
    current_entity = None
    current_layer = "0"

    for i, line in enumerate(lines):
        stripped = line.strip()
        next_line = lines[i + 1].strip() if i + 1 < len(lines) else ""

        if stripped == "0" and next_line:
            if current_entity:
                entities.append(current_entity)
            current_entity = {"type": next_line, "layer": current_layer}
        elif stripped == "8" and next_line:
            current_layer = next_line
            layers.add(next_line)
            if current_entity:
                current_entity["layer"] = next_line

    if current_entity:
        entities.append(current_entity)

    entity_types = list(set(e["type"] for e in entities if e.get("type")))

    return {
        "ext": "dxf",
        "text": f"DXF file with {len(entities)} entities across {len(layers)} layers. "
                f"Layers: {', '.join(sorted(layers))}. "
                f"Entity types: {', '.join(entity_types)}.",
        "metadata": {
            "layers": sorted(layers),
            "entity_types": entity_types,
            "entity_count": len(entities),
        },
    }


def _parse_geojson(content: bytes) -> dict:
    """Parse GeoJSON files."""
    import json
    data = json.loads(content)
    features = data.get("features", [])
    types = list(set(f.get("geometry", {}).get("type") for f in features if f.get("geometry")))
    props = []
    if features:
        first_props = features[0].get("properties", {}) or {}
        props = list(first_props.keys())[:20]

    return {
        "ext": "geojson",
        "text": f"GeoJSON with {len(features)} features. "
                f"Types: {', '.join(types)}. "
                f"Properties: {', '.join(props)}.",
        "metadata": {
            "feature_count": len(features),
            "geometry_types": types,
            "properties": props,
        },
    }


def _parse_image(content: bytes) -> dict:
    """Extract text from images using OCR (pytesseract)."""
    try:
        from PIL import Image
        import pytesseract
        img = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(img, lang="eng")
        return {
            "ext": "img",
            "text": text,
            "metadata": {
                "ocr": True,
                "width": img.width,
                "height": img.height,
                "format": img.format,
            },
        }
    except ImportError:
        return {
            "ext": "img",
            "text": "[OCR not available: pytesseract not installed]",
            "metadata": {"ocr": False},
        }
    except Exception as e:
        return {
            "ext": "img",
            "text": f"[OCR error: {e}]",
            "metadata": {"ocr": False, "error": str(e)},
        }


def _parse_text(content: bytes) -> dict:
    """Parse plain text files (TXT, RAW, GSI, etc.)."""
    text = content.decode("utf-8", errors="replace")
    lines = text.strip().split("\n")
    return {
        "ext": "txt",
        "text": text,
        "metadata": {"line_count": len(lines), "char_count": len(text)},
    }